from buho_back.config import TEMPLATES_DIRECTORY
from buho_back.utils import safe_cast
import os
from datetime import datetime
import openpyxl

templates_directory = TEMPLATES_DIRECTORY


def find_best_scale(number):
    # Define the scales
    scales = [1000, 1000000, 1000000000]  # thousands, millions, billions
    scale_units = ["K", "Mi", "Bi"]

    # Find the appropriate scale
    for scale, unit in zip(scales, scale_units):
        if number / scale < 999999:
            return scale, unit

    # If the number is very large, return the largest scale
    return scales[-1], scale_units[-1]


def generate_dcf(input_variables, template_path, user_output_files_directory, filename):
    print(input_variables)
    scale, scale_units = find_best_scale(input_variables["EBIT"])
    input_variables["Scale"] = scale_units
    input_variables["EBIT"] = input_variables["EBIT"]/scale
    input_variables["D&A"] = input_variables["D&A"]/scale
    input_variables["EBITDA"] = input_variables["EBIT"]+input_variables["D&A"]
    input_variables["D&A (%EBITDA)"] = input_variables["D&A"]/input_variables["EBITDA"]

    # Open the existing Excel file
    workbook = openpyxl.load_workbook(template_path)

    # Select the worksheet named "input variables"
    if "input variables" not in workbook.sheetnames:
        raise ValueError(
            "The sheet named 'input variables' does not exist in the provided Excel file."
        )
    sheet = workbook["input variables"]

    # Iterate over the rows in column A to match dict keys
    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=False):
        cell_key = row[0]
        cell_value = row[1]
        if cell_key.value in input_variables:
            cell_value.value = input_variables[cell_key.value]

    # Define the new file path
    if not os.path.exists(user_output_files_directory):
        os.makedirs(user_output_files_directory)
    new_file_path = os.path.join(f"{user_output_files_directory}", f"{filename}.xlsx")

    # Save the modified workbook to the new path
    workbook.save(new_file_path)

    return new_file_path


def generate_xlsx(content, user_output_files_directory, filename, user_parameters):
    template_path = os.path.join(templates_directory, f"{filename}.xlsx")
    ## get template, fill variables with content, and save as new output file
    if filename == "discounted_cash_flow":
        user_parameters = {
            key: safe_cast(value) for key, value in user_parameters.items()
        }
        dcf_parameters = user_parameters | content
        return generate_dcf(
            dcf_parameters, template_path, user_output_files_directory, filename
        )
